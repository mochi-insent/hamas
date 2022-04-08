#   ■基本測定データの判定ver2(20220311)対応
# #   ラベルプリンタPC から
# ファイル変更イベント検出のため、watchdogをインポート
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

# ファイルアクセスとスリープのため、osとtimeをインポート
import os
import time
import pandas as pd
import openpyxl

def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

# 監視対象ディレクトリを指定する
target_dir = '\\\\192.168.24.27\\disk1\\New共通\\生産部\\品質保証\\05_生産\\02_生産管理\\02_工程管理\\測定値記録自動化\\濱田さんEXCEL\\'
dst_file = '★基本測定データの判定.xlsm'

Ver = 'ver_20211224'
#   Ver = 'ver_20220311'

if Ver == 'ver_20211224':
    ref_file = '■基本測定データの判定(20211224) - コピー.xlsm'
    h_sheet_name = '7ｻﾝﾌﾟﾙ3周(Z)'               #   ■基本測定データの判定ver2(20220311)対応
elif Ver == 'ver_20220311':
    ref_file = '■基本測定データの判定ver2(20220311) - コピー.xlsm'
    h_sheet_name = '7ｻﾝﾌﾟﾙ3周'               #   ■基本測定データの判定ver2(20220311)対応   

# FileSystemEventHandler の継承クラスを作成
class FileChangeHandler(FileSystemEventHandler):
     # ファイル作成時のイベント
     def on_created(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         print('%s created' % filename)
         #   リストに列名をつけて取り込んだ列長がデコボコにならないようにする
         col_name = ['c{0:02d}'.format(i) for i in range(20)]
         if filename[-7:-4] == 'prt':
             if os.path.exists(target_dir + dst_file):
                 #   既にある出力ファイルを読み込む
                 hamadabook = openpyxl.load_workbook(target_dir + dst_file, keep_vba=True)
             else:
                 #   濱田さん判定excelを取り込む
                 hamadabook = openpyxl.load_workbook(target_dir + ref_file, keep_vba=True)
             #   prt読み込み実行
             prt_values = pd.read_csv(filepath, header = None, encoding = "shift-jis", names = col_name, skip_blank_lines=False)
             #   必要なデータ部分を切り取る
             prt_values_s = prt_values.iloc[23:78, 0:17]
 
             #
             #   prtから切り出す範囲がフレキシブルであるときの処理、'センサー間の相関係数'の手前まで読む。固定範囲のときはprt_values_s が単純にprt_values_ss　にコピーされる。
             #
             prt_values_ss = []
             data = []
             for row, data in prt_values_s.iterrows():
                 if data[0] == 'センサー間の相関係数':
                     break
                 else:
                     prt_values_ss.append(data)

             #   貼り付け先シート（固定！）　　★他のシートにも拡張必要
             h_sheet = hamadabook[h_sheet_name]

             #   prtデータを貼り付け（行、列　＝　４，３から）
             write_list_2d(h_sheet, prt_values_ss, 4, 3)

             #   excelファイル保存
             hamadabook.save(target_dir + dst_file)

         elif   filename[-7:-4] == 'abs':
             if os.path.exists(target_dir + dst_file):
                 #   既にある出力ファイルを読み込む
                 hamadabook = openpyxl.load_workbook(target_dir + dst_file, keep_vba=True)
             else:
                 #   濱田さん判定excelを取り込む
                 hamadabook = openpyxl.load_workbook(target_dir + ref_file, keep_vba=True)
             #   abs読み込み実行
             abs_values = pd.read_csv(filepath, header=None, encoding = "shift-jis", names = col_name, skip_blank_lines=False)
 
             #   必要なデータ部分を切り取る
             abs_values_s = abs_values.iloc[:, 3:14]
 
             #
             #   absから切り出す範囲がフレキシブルであるときの処理、空欄（Nan）が見つかるまで読む。固定範囲のときはprt_values_s が単純にprt_values_ss　にコピーされる。
             #
             abs_values_ss = []
             for row, data in abs_values_s.iterrows():
                 if data.hasnans:
                     break
                 else:
                     abs_values_ss.append(data)
 
             #   貼り付け先シート（固定！）　　★他のシートにも拡張必要
             h_sheet = hamadabook[h_sheet_name]

             #   absデータを貼り付け（行、列　＝　６５，２から）
             write_list_2d(h_sheet, abs_values_ss, 65, 2)

             #   excelファイル保存
             hamadabook.save(target_dir + dst_file)


     # ファイル変更時のイベント
     def on_modified(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         print('%s changed' % filename)

     # ファイル削除時のイベント
     def on_deleted(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         print('%s deleted' % filename)

     # ファイル移動時のイベント
     def on_moved(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         print('%s moved' % filename)

# コマンド実行の確認
if __name__ == "__main__":
     # ファイル監視の開始
     event_handler = FileChangeHandler()
     observer = Observer()
     observer.schedule(event_handler, target_dir, recursive=True)
     observer.start()
     # 処理が終了しないようスリープを挟んで無限ループ
     try:
         while True:
             time.sleep(0.1)
     except KeyboardInterrupt:
         observer.stop()
     observer.join()
     