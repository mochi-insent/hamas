#   ###### センサ基本測定データ判定ファイル名.txt　で読み込み元ファイル名を指定する　#######
# #   ラベルプリンタPC から

#   ロギングの設定（jsonファイルから）
import json
from logging import getLogger, config
from this import d

with open('./log_config.json', 'r') as f:
    log_conf = json.load(f)

config.dictConfig(log_conf)

logger = getLogger(__name__)

#   設定ファイルから読込
from hamadasConfigure import hamasconf

# ファイル変更イベント検出のため、watchdogをインポート
import re
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

# ファイルアクセスとスリープのため、osとtimeをインポート
import os
import time
import openpyxl
import csv

#   矩形コピペ、リストl_2d（２次元）を指定シートの指定位置から貼る
def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

#   浮動小数点の文字列を数値に変換。浮動小数点の格好をしていないものはそのまま返す。
def to_float(s):
    try:
        return float(s)
    except ValueError:
        return s


# 監視対象ディレクトリ
target_dir = hamasconf.target_dir
logger.info(target_dir)

# 参照ファイル名
f = hamasconf.f
ref_file = f.read()
logger.info(f)

#   出力ファイル名
dst_file = hamasconf.dst_file
logger.info(dst_file)

#   出力シート名、貼付け開始位置
h_sheet_name = hamasconf.h_sheet_name
start_row_prt = hamasconf.start_row_prt
start_col_prt = hamasconf.start_col_prt
start_row_abs = hamasconf.start_row_abs
start_col_abs = hamasconf.start_col_abs

#   ハートビート用タイムスタンプの書き込み先
path = hamasconf.path
logger.info(path)

# FileSystemEventHandler の継承クラスを作成
class FileChangeHandler(FileSystemEventHandler):
     # ファイル作成時のイベント
     def on_created(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         dst_dir = filepath[:(filepath.rfind('\\')+1)]
         logger.info('%s %s created' % (filename,  str(dst_dir)))

         #   abs、prt いずれも処理が終わったかどうかを調べるフラグ
         flg_rename = 0

         #  ｐｒｔファイル処理
         if filename[-7:-4] == 'prt':
             dst_file = '★'+ filename[:-8] + '.xlsm'
             if os.path.exists(dst_dir + dst_file):
                 #   既にある出力ファイルを読み込む、あとで結果ファイルをリネームするフラグを立てる
                 hamadabook = openpyxl.load_workbook(dst_dir + dst_file, keep_vba=True)
                 flg_rename = 1
             else:
                 #   濱田さん判定excelを取り込む
                 hamadabook = openpyxl.load_workbook(ref_file, keep_vba=True)
             #  ファイルを２次元リストとして取り込む（イテレータ反復処理）
             with open(filepath,newline="") as csvf:
                 prt_values=csv.reader(csvf)
                 prt_values_s = [row for row in prt_values]

             #   必要なデータ部分を切り取る
                 prt_values_ss = [r[0:17] for r in prt_values_s]
            
             #  数字文字列を数値（浮動小数点）に変換
             prt_values_sf = [[to_float(ss) for ss in s] for s in prt_values_ss]
 
             #   貼り付け先シート（固定！）　　
             h_sheet = hamadabook[h_sheet_name]

             #   prtデータを貼り付け（行、列　＝　４，３から）
             write_list_2d(h_sheet, prt_values_sf, start_row=start_row_prt, start_col=start_col_prt)

             #   excelファイル保存
             hamadabook.save(dst_dir + dst_file)
             hamadabook.close()

             if flg_rename:
                 os.rename(dst_dir + dst_file, dst_dir + '★' + dst_file)

         #  ａｂｓファイル処理
         elif   filename[-7:-4] == 'abs':
             dst_file = '★'+ filename[:-8] + '.xlsm'
             if os.path.exists(dst_dir + dst_file):
                 #   既にある出力ファイルを読み込む、あとで結果ファイルをリネームするフラグを立てる
                 hamadabook = openpyxl.load_workbook(dst_dir + dst_file, keep_vba=True)
                 flg_rename = 1
             else:
                 #   濱田さん判定excelを取り込む
                 hamadabook = openpyxl.load_workbook(ref_file, keep_vba=True)
             #   ファイルを２次元リストとして取り込む（イテレータ反復処理）
             with open(filepath,newline="") as csvf:
                 abs_values = csv.reader(csvf)
                 abs_values_s = [row for row in abs_values]
 
             #   必要なデータ部分を切り取る
             try:
                col = abs_values_s[0].index('温度センサー')
             #  温度センサーがない場合はＭ列まで取り込む
             except:
                col = 13
             abs_values_ss = [r[0:col] for r in abs_values_s]

             #  数字文字列を数値（浮動小数点）に変換
             abs_values_sf = [[to_float(ss) for ss in s] for s in abs_values_ss]
 
             #   貼り付け先シート（固定！）
             h_sheet = hamadabook[h_sheet_name]

             #   absデータを貼り付け（行、列　＝　６５，２から）
             write_list_2d(h_sheet, abs_values_sf, start_row=start_row_abs, start_col=start_col_abs)

             #   excelファイル保存
             hamadabook.save(dst_dir + dst_file)
             hamadabook.close()

             if flg_rename:
                 os.rename(dst_dir + dst_file, dst_dir + '★' + dst_file)



     # ファイル変更時のイベント
     def on_modified(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         logger.info('%s changed' % filename)

     # ファイル削除時のイベント
     def on_deleted(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         logger.info('%s deleted' % filename)

     # ファイル移動時のイベント
     def on_moved(self, event):
         filepath = event.src_path
         filename = os.path.basename(filepath)
         logger.info('%s moved' % filename)

# コマンド実行の確認
if __name__ == "__main__":
     # ファイル監視の開始
     logger.info('Process Started.....')
     event_handler = FileChangeHandler()
     observer = Observer()
     observer.schedule(event_handler, target_dir, recursive=True)
     observer.start()
     # 処理が終了しないようスリープを挟んで無限ループ
     try:
         while True:
             time.sleep(0.2)
             u = str(time.time())
             f = open(path, 'w')
             f.write(u)
             f.close
     except KeyboardInterrupt:
         observer.stop()
     observer.join()
     